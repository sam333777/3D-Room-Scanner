import numpy as np
import open3d as o3d
from openpyxl import load_workbook

if __name__ == "__main__":
    workbook = load_workbook(filename='data.xlsx')  # Load Excel workbook
    worksheet = workbook.active  # Get the active worksheet

    # Convert Excel data to numpy array
    data = []
    for row_cells in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, max_col=3, values_only=True):
        data.append(row_cells)

    # Convert data to numpy array
    data = np.array(data)

    # Group points based on their X coordinates
    x_groups = {}
    for point in data:
        x = int(point[0])  # Assuming X displacement is integer
        if x not in x_groups:
            x_groups[x] = []
        x_groups[x].append(point)

    # Create LineSets for each X displacement plane
    line_sets = []
    for x, points in x_groups.items():
        lines = []
        num_points = len(points)
        for i in range(num_points - 1):
            lines.append([i, i + 1])
        line_set = o3d.geometry.LineSet(
            points=o3d.utility.Vector3dVector(points),
            lines=o3d.utility.Vector2iVector(lines)
        )
        line_sets.append(line_set)

    # Visualize the point clouds and lines
    o3d.visualization.draw_geometries([*line_sets])
