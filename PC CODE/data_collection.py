import serial
import math
from openpyxl import Workbook

ser = serial.Serial('/dev/cu.usbmodemME4010231', 115200)  # Change COM Port #

workbook = Workbook()  # Create Excel workbook
worksheet = workbook.active  # Get the active worksheet

i = 0
x = 0  # Change initial x-displacement (mm)
increment = 200  # x-displacement increment (mm)

while True:
    s = ser.readline().decode("utf-8").strip()  # Read and decode UART data
    data = s.split(",")  # Split the data by comma
    
    if len(data) >= 2:  # Check if the second element is a digit (distance in mm)
        a = int(data[1])  # Convert distance to an integer
        angle = (i / 512) * 2 * math.pi  # Obtain angle based on motor rotation
        y = a * math.cos(angle)  # Calculate y
        z = a * math.sin(angle)  # Calculate z
        x = int(data[3]) * 200
        print(x, y, z)
        # Write x, y, z values to the Excel file
        worksheet.cell(row=i+1, column=1, value=x)  # Write x value to column A
        worksheet.cell(row=i+1, column=2, value=y)  # Write y value to column B
        worksheet.cell(row=i+1, column=3, value=z)  # Write z value to column C
        
        i += 1
        
          # Increment x position
            
    elif not s:  # If the line is empty (end of data)
        break
    print(data)
    workbook.save("data.xlsx")  # Save the workbook
